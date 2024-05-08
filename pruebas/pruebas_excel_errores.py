# -*- coding: utf-8 -*-
"""
Created on Tue May  7 17:05:05 2024

@author: jeconchao
"""
from openpyxl import load_workbook, Workbook
from pathlib import Path



def anotar_error(pregunta, error):
    if not Path('problemas_datos.xlsx').is_file():
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Pregunta'
        ws['B1'] = 'Error'
        wb.save('problemas_datos.xlsx')
    
    wb = load_workbook(filename='problemas_datos.xlsx')
    ws = wb.active
    ws.append([pregunta, error])
    wb.save('problemas_datos.xlsx')
