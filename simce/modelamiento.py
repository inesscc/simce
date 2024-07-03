import torch.nn as nn
from pathlib import Path
from openpyxl import load_workbook, Workbook


def preparar_capas_modelo(model, modelo_seleccionado):
    num_classes = 2
    print(f'{modelo_seleccionado=}')
    if modelo_seleccionado == 'vgg16':           
        num_features = model.classifier[6].in_features
        model.classifier[6] = nn.Linear(num_features, num_classes)
    elif modelo_seleccionado == 'maxvit_t':
        num_features = model.classifier[5].in_features
        model.classifier[5] = nn.Linear(num_features, num_classes)
    elif modelo_seleccionado == 'wide_resnet101_2':
        num_features = model.fc.in_features
        model.fc = nn.Linear(num_features, num_classes)
    elif 'efficientnet' in modelo_seleccionado:
        num_features = model.classifier[1].in_features
        model.classifier[1] = nn.Linear(num_features, num_classes)

    else:
        raise('Código no está preparado para incorporar modelo. Modifique función para incorporarla')
    
    return model

def anotar_metricas_modelo(config_name, run_id, log):
    if not Path('metricas_modelos.xlsx').is_file():
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'ID_modelo'
        ws['B1'] = 'val_loss'
        ws['C1'] = 'val_acc'
        wb.save('metricas_modelos.xlsx')

    id_modelo = f'{config_name}-{run_id}'

    wb = load_workbook(filename='metricas_modelos.xlsx')

    ws = wb.active
    id_modelos = {cell[0].value for cell in ws.iter_rows(min_col=1, max_col=1)}

    if id_modelo not in id_modelos:
        ws.append([id_modelo, log['loss'], log['accuracy']])
        wb.save('metricas_modelos.xlsx')
        print('Métricas guardadas exitosamente!')
    else:
        print('MODELO YA ANOTADO ANTERIORMENTE, no fue anotado.')