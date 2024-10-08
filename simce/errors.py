# -*- coding: utf-8 -*-
"""
Created on Thu May  2 16:17:41 2024

@author: jeconchao
"""

from openpyxl import load_workbook, Workbook
from pathlib import Path
from simce.utils import timing

def anotar_error(pregunta, error, nivel_error, e=None):

    print(error)

    if e:
        print(e)

    if not Path('problemas_imagenes.xlsx').is_file():
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Pregunta'
        ws['B1'] = 'Error'
        ws['C1'] = 'Nivel'
        wb.save('problemas_imagenes.xlsx')

    wb = load_workbook(filename='problemas_imagenes.xlsx')

    ws = wb.active

    pregs_con_error = {cell[0].value for cell in ws.iter_rows(min_col=1, max_col=1)}

    if pregunta not in pregs_con_error:
        print('ANOTANDO ERROR -----')
        ws.append([pregunta, error, nivel_error])
        wb.save('problemas_imagenes.xlsx')
    else:
        print('ERROR YA ANOTADO ANTERIORMENTE, no fue anotado.')


def agregar_error(queue, pregunta, error, nivel_error):
    """agrega la dupla a la fila para añadir el error al finalizar el multi-procesamiento"""
    queue.put((pregunta, error, nivel_error))

@timing
def escribir_errores(queue):
    """une todos los errores generados en las iteraciones de los diferentes procesos"""
    if not Path('problemas_imagenes.xlsx').is_file():
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Pregunta'
        ws['B1'] = 'Error'
        ws['C1'] = 'Nivel'
        wb.save('problemas_imagenes.xlsx')

    wb = load_workbook(filename='problemas_imagenes.xlsx')
    ws = wb.active

    while not queue.empty():
        pregunta, error, nivel_error = queue.get()
        ws.append([pregunta, error, nivel_error])

    wb.save('problemas_imagenes.xlsx')
    
